"""Carga masiva por Excel/CSV: productos, movimientos y catálogo concepto → clave SAT."""

import io
from pathlib import Path

import openpyxl

from tests.conftest import agregar_membresia, auth_headers, crear_empresa, crear_usuario

FIXTURES = Path(__file__).parent / "fixtures"


def _xlsx(filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    for f in filas:
        ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup(client, db, seed_rbac):
    usuario = crear_usuario(db)
    empresa = crear_empresa(db)
    agregar_membresia(db, usuario=usuario, empresa=empresa, rol=seed_rbac["administrador"])
    return auth_headers(client, email=usuario.email, password="Demo1234!", empresa_id=empresa.id), empresa


def _subir(client, headers, ruta, contenido, nombre):
    return client.post(ruta, headers=headers, files={"archivo": (nombre, contenido, "application/octet-stream")})


def test_plantillas_descargables(client, seed_rbac, db):
    headers, _ = _setup(client, db, seed_rbac)
    for ruta in ("/api/v1/inventory/productos/plantilla", "/api/v1/inventory/movimientos/plantilla", "/api/v1/pedimentos/conceptos/plantilla"):
        res = client.get(ruta, headers=headers)
        assert res.status_code == 200 and res.content[:2] == b"PK"
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        assert wb.active.max_row >= 2 and "Instrucciones" in wb.sheetnames


def test_carga_productos_crea_actualiza_y_reporta_errores(client, seed_rbac, db):
    headers, _ = _setup(client, db, seed_rbac)
    archivo = _xlsx(
        [
            ["Catálogo de productos"],  # título antes del encabezado
            ["SKU", "Nombre", "Tipo", "Categoría", "Unidad SAT", "Costo unitario", "Clave SAT", "Color", "Talla"],
            ["CAL-001", "Calzado dama", "producto", "Calzado", "PR", "674.11", "53111600", "negro", "25"],
            ["SRV-01", "Flete local", "servicio", "Logística", "E48", 1500, None, None, None],
            ["CAL-002", "Calzado caballero", "raro", "Calzado", "PR", 700, None, None, None],  # tipo inválido
            ["CAL-003", "", "producto", None, None, 1, None, None, None],  # sin nombre
            ["CAL-001", "Duplicado en archivo", "producto", None, None, 1, None, None, None],
            ["CAL-004", "Clave mala", "producto", None, None, 1, "123", None, None],
        ]
    )
    res = _subir(client, headers, "/api/v1/inventory/productos/importar", archivo, "productos.xlsx")
    assert res.status_code == 200, res.text
    r = res.json()
    assert (r["creados"], r["actualizados"]) == (2, 0)
    assert {e["fila"] for e in r["errores"]} == {5, 6, 7, 8}
    productos = {p["sku"]: p for p in client.get("/api/v1/inventory/productos", headers=headers).json()}
    assert productos["CAL-001"]["atributos"] == {"Color": "negro", "Talla": "25"}
    assert productos["CAL-001"]["clave_prodserv"] == "53111600" and productos["CAL-001"]["costo_unitario"] == 674.11
    assert productos["SRV-01"]["tipo"] == "servicio"

    # Segunda carga: mismo SKU actualiza (costo y nombre), CSV
    csv = "sku,nombre,costo\nCAL-001,Calzado dama premium,800\nNEW-1,Nuevo,10\n".encode()
    r2 = _subir(client, headers, "/api/v1/inventory/productos/importar", csv, "p.csv").json()
    assert (r2["creados"], r2["actualizados"], r2["errores"]) == (1, 1, [])
    p = client.get("/api/v1/inventory/productos", headers=headers).json()
    assert next(x for x in p if x["sku"] == "CAL-001")["costo_unitario"] == 800

    # Sin columnas mínimas → 422
    assert _subir(client, headers, "/api/v1/inventory/productos/importar", b"a,b\n1,2\n", "x.csv").status_code == 422


def test_carga_movimientos_aplica_validas_y_reporta_invalidas(client, seed_rbac, db):
    headers, _ = _setup(client, db, seed_rbac)
    client.post("/api/v1/inventory/almacenes", headers=headers, json={"nombre": "Central", "codigo": "CEN"})
    client.post("/api/v1/inventory/productos", headers=headers, json={"sku": "A1", "nombre": "Prod A", "costo_unitario": 10})
    archivo = _xlsx(
        [
            ["SKU", "Almacén", "Tipo", "Cantidad", "Costo unitario", "Referencia", "Nota"],
            ["A1", "CEN", "entrada", 100, 12.5, "OC-1", "inicial"],
            ["A1", "CEN", "salida", 30, None, "REM-1", None],
            ["A1", "CEN", "salida", 500, None, "REM-2", None],  # stock insuficiente
            ["ZZ", "CEN", "entrada", 5, None, None, None],  # producto inexistente
            ["A1", "NOPE", "entrada", 5, None, None, None],  # almacén inexistente
            ["A1", "CEN", "ajuste", -20, None, "AJ-1", "merma"],
            ["A1", "CEN", "entrada", 2.5, None, None, None],  # decimal
        ]
    )
    r = _subir(client, headers, "/api/v1/inventory/movimientos/importar", archivo, "movs.xlsx").json()
    assert r["creados"] == 3
    assert {e["fila"] for e in r["errores"]} == {4, 5, 6, 8}
    stock = client.get("/api/v1/inventory/stock", headers=headers).json()
    assert stock[0]["disponible"] == 50  # 100 − 30 − 20
    movs = client.get("/api/v1/inventory/movimientos", headers=headers).json()
    entrada = next(m for m in movs if m["referencia"] == "OC-1")
    assert entrada["costo_unitario"] == 12.5


def test_catalogo_conceptos_desde_excel_del_cliente_y_uso_en_pedimentos(client, seed_rbac, db):
    headers, _ = _setup(client, db, seed_rbac)
    contenido = (FIXTURES / "catalogo_conceptos_sat.xlsx").read_bytes()
    r = _subir(client, headers, "/api/v1/pedimentos/conceptos/importar", contenido, "CATALOGO.xlsx").json()
    assert r["creados"] > 1000 and r["total_catalogo"] == r["creados"]
    # Reimportar: todo actualizados, nada nuevo
    r2 = _subir(client, headers, "/api/v1/pedimentos/conceptos/importar", contenido, "CATALOGO.xlsx").json()
    assert r2["creados"] == 0 and r2["actualizados"] == r["creados"]
    lista = client.get("/api/v1/pedimentos/conceptos?q=acumuladores", headers=headers).json()
    assert any(c["concepto"].upper() == "ACUMULADORES" and c["clave_prodserv"] == "26111700" for c in lista["items"])

    # Al importar el M3, las partidas cuyo concepto está en el catálogo reciben clave
    with (FIXTURES / "m3382475.003").open("rb") as f:
        ped = client.post("/api/v1/pedimentos/importar", headers=headers, files={"archivo": ("m.003", f, "application/octet-stream")}).json()["pedimento"]
    con_clave = [p for p in ped["partidas"] if p["clave_prodserv"]]
    assert con_clave, "ninguna partida recibió clave del catálogo"
    # Alta manual de un concepto y re-aplicar al pedimento
    sin = next(p for p in ped["partidas"] if not p["clave_prodserv"])
    client.put("/api/v1/pedimentos/conceptos", headers=headers, json={"concepto": sin["descripcion"], "clave_prodserv": "10101500"})
    ped2 = client.post(f"/api/v1/pedimentos/{ped['id']}/aplicar-claves", headers=headers).json()
    assert next(p for p in ped2["partidas"] if p["id"] == sin["id"])["clave_prodserv"] == "10101500"
    assert client.put("/api/v1/pedimentos/conceptos", headers=headers, json={"concepto": "X", "clave_prodserv": "12"}).status_code == 422
