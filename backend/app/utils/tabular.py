"""Lectura tolerante de archivos tabulares (xlsx / csv) para las cargas masivas.

Todas las importaciones (estados de cuenta, productos, movimientos, catálogos)
comparten esta capa: se lee la primera hoja como matriz, se buscan encabezados
por sinónimos (sin acentos, minúsculas) y se normalizan números y fechas.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


class ArchivoTabularError(ValueError):
    pass


def norm(texto: object) -> str:
    t = unicodedata.normalize("NFKD", str(texto or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", t.strip().lower())


def a_decimal(valor: object) -> Decimal | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    t = str(valor).strip()
    if not t or t in ("-", "—"):
        return None
    negativo = t.startswith("(") and t.endswith(")")
    t = re.sub(r"[^\d,.\-]", "", t)
    if t.count(",") and t.count("."):
        t = t.replace(",", "")
    elif t.count(",") == 1 and len(t.split(",")[1]) in (1, 2):
        t = t.replace(",", ".")  # 1234,56
    else:
        t = t.replace(",", "")
    try:
        d = Decimal(t)
    except InvalidOperation:
        return None
    return -d if negativo else d


_FORMATOS_FECHA = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y", "%d/%b/%Y", "%d-%b-%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S")


def a_fecha(valor: object) -> date | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    t = str(valor).strip()
    if not t:
        return None
    for f in _FORMATOS_FECHA:
        try:
            return datetime.strptime(t, f).date()
        except ValueError:
            continue
    return None


def leer_tabla(contenido: bytes, nombre: str) -> list[list[object]]:
    """Devuelve la primera hoja (xlsx) o el CSV como lista de filas."""
    nombre = (nombre or "").lower()
    if nombre.endswith((".xlsx", ".xlsm")) or contenido[:2] == b"PK":
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise ArchivoTabularError("Falta la librería openpyxl para leer Excel") from exc
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    texto = None
    for enc in ("utf-8-sig", "latin-1"):
        try:
            texto = contenido.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ArchivoTabularError("No se pudo decodificar el archivo de texto")
    try:
        dialecto = csv.Sniffer().sniff(texto[:4096], delimiters=",;\t|")
    except csv.Error:
        dialecto = csv.excel
    return [row for row in csv.reader(io.StringIO(texto), dialecto)]


def detectar_columnas(encabezados: list[object], sinonimos: dict[str, list[str]]) -> dict[str, int]:
    """Mapea campo → índice de columna buscando cada sinónimo (igual o prefijo)."""
    normal = [norm(h) for h in encabezados]
    mapa: dict[str, int] = {}
    for campo, lista in sinonimos.items():
        for i, h in enumerate(normal):
            if not h or i in mapa.values():
                continue
            if h in lista or any(h.startswith(s) for s in lista):
                mapa[campo] = i
                break
    return mapa


def localizar_encabezado(tabla: list[list[object]], sinonimos: dict[str, list[str]], requeridos: set[str], max_filas: int = 30) -> tuple[int, dict[str, int]]:
    """Busca en las primeras filas la que contenga todas las columnas requeridas."""
    for i, fila in enumerate(tabla[:max_filas]):
        m = detectar_columnas(fila, sinonimos)
        if requeridos <= set(m):
            return i, m
    raise ArchivoTabularError(
        "No se encontró un encabezado con las columnas requeridas: " + ", ".join(sorted(requeridos))
    )


def celda(fila: list[object], mapa: dict[str, int], campo: str) -> object:
    idx = mapa.get(campo)
    if idx is None or idx >= len(fila):
        return None
    v = fila[idx]
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def texto(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def plantilla_xlsx(hoja: str, columnas: list[str], ejemplo: list[object] | None = None, notas: list[str] | None = None) -> bytes:
    """Genera un .xlsx con encabezados (y una fila de ejemplo) para que el usuario lo llene."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja[:31]
    ws.append(columnas)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
    if ejemplo:
        ws.append(ejemplo)
    for i, col in enumerate(columnas, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, min(40, len(col) + 4))
    if notas:
        ws2 = wb.create_sheet("Instrucciones")
        for n in notas:
            ws2.append([n])
        ws2.column_dimensions["A"].width = 110
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
